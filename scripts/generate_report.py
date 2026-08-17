#!/usr/bin/env python3
"""Generate the client-facing Excel scouting report from the database.

Two data sources:

  --source render   (default)  Pull live rows from the Render Postgres via the
                               Render CLI (`render psql … COPY … TO STDOUT`).
                               Needs: `render` CLI on PATH, plus RENDER_API_KEY
                               and SCOUTING_DB_ID (from the environment or
                               ~/.render_env; --db-id overrides the latter).

  --source csv --csv-dir DIR   Build from a folder of CSV exports produced by
                               this script or by a manual COPY (one file per
                               table: sessions.csv, prospects.csv,
                               observations.csv, scout_profiles.csv).

Output: reports/informe-scouting-<date>.xlsx (5 sheets: Resumen, Partidos,
Jugadores, Observaciones, Notas equipo). The whole DB is included — filter to a
single Telegram account with --chat-id if needed.

Dependency: openpyxl (pip install openpyxl).

Examples:
    source ~/.render_env && python scripts/generate_report.py
    python scripts/generate_report.py --source csv --csv-dir reports/db-export-limpia-2026-07-30
"""
from __future__ import annotations

import argparse
import csv
import io
import os
import subprocess
import sys
from collections import defaultdict
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    sys.exit("openpyxl is required — install it with:  pip install openpyxl")

# Which Render Postgres to read. Kept out of the source because this repo is
# public and the id names the production database: set SCOUTING_DB_ID in the
# environment (~/.render_env alongside RENDER_API_KEY is the usual place) or
# pass --db-id.
DB_ID_VAR = "SCOUTING_DB_ID"
TABLES = ("sessions", "scout_profiles", "prospects", "observations")

# Manual 1–5 rating → decision label (mirrors scouting_bot/models.py).
RATING_DECISIONS = {1: "A descartar", 2: "A seguir", 3: "Interesante",
                    4: "Muy interesante", 5: "A firmar"}


# ── data loading ──────────────────────────────────────────────────────────
def _load_render_env() -> None:
    """Populate RENDER_API_KEY and SCOUTING_DB_ID from ~/.render_env.

    Values already in the environment win, so `source ~/.render_env` and a
    plain run behave the same."""
    path = os.path.expanduser("~/.render_env")
    if not os.path.exists(path):
        return
    wanted = ("RENDER_API_KEY", DB_ID_VAR)
    with open(path) as fh:
        for line in fh:
            line = line.strip().removeprefix("export ").strip()
            key, sep, value = line.partition("=")
            if sep and key in wanted and not os.environ.get(key):
                os.environ[key] = value.strip().strip("'\"")


def load_from_render(db_id: str | None) -> dict[str, list[dict]]:
    _load_render_env()
    if not os.environ.get("RENDER_API_KEY"):
        sys.exit("No RENDER_API_KEY. Run `source ~/.render_env` first, or use --source csv.")
    db_id = db_id or os.environ.get(DB_ID_VAR)
    if not db_id:
        sys.exit(
            f"No database id. Pass --db-id, or set {DB_ID_VAR} "
            "(in the environment or ~/.render_env)."
        )
    data: dict[str, list[dict]] = {}
    for t in TABLES:
        cmd = ["render", "psql", db_id, "--confirm", "--command",
               f"COPY (SELECT * FROM {t} ORDER BY 1) TO STDOUT WITH CSV HEADER"]
        out = subprocess.run(cmd, capture_output=True, text=True)
        if out.returncode != 0:
            sys.exit(f"render psql failed for {t}:\n{out.stderr.strip()}")
        data[t] = list(csv.DictReader(io.StringIO(out.stdout)))
    return data


def load_from_csv(csv_dir: str) -> dict[str, list[dict]]:
    data: dict[str, list[dict]] = {}
    for t in TABLES:
        path = os.path.join(csv_dir, f"{t}.csv")
        if not os.path.exists(path):
            sys.exit(f"Missing {path}")
        with open(path, newline="") as fh:
            data[t] = list(csv.DictReader(fh))
    return data


# ── value helpers ─────────────────────────────────────────────────────────
_NULLS = ("", None, r"\N")


def _int(v):
    return int(v) if v not in _NULLS else None


def _float(v):
    try:
        return float(v) if v not in _NULLS else None
    except ValueError:
        return None


def _bool(v) -> bool:
    return str(v).lower() in ("t", "true", "1")


def decision_for_rating(rating) -> str | None:
    r = _float(rating)
    if r is None:
        return None
    return RATING_DECISIONS[min(5, max(1, round(r)))]


def _clean(text) -> str:
    return (text or "").replace("\n", " ").replace("\r", " ").strip()


# ── styling ───────────────────────────────────────────────────────────────
GREEN = "14532D"
HDR_FILL = PatternFill("solid", fgColor="EAF4EE")
HDR_FONT = Font(bold=True, color=GREEN)
_thin = Side(style="thin", color="D8E6DD")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def _style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER


def _widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# ── report builder ────────────────────────────────────────────────────────
def build(data: dict[str, list[dict]], out_path: str, report_date: str) -> dict:
    sessions = data["sessions"]
    prospects = data["prospects"]
    observations = data["observations"]
    scouts = data["scout_profiles"]

    sess_by_id = {s["id"]: s for s in sessions}

    def match_label(sid):
        s = sess_by_id.get(sid)
        return f"{s['home_team']} vs {s['away_team']}" if s else "?"

    obs_by_prosp: dict[str, list[dict]] = defaultdict(list)
    team_notes: list[dict] = []
    for o in observations:
        if _bool(o["is_team_note"]):
            team_notes.append(o)
        elif o["prospect_id"] not in _NULLS:
            obs_by_prosp[o["prospect_id"]].append(o)

    def obs_line(o):
        q = _clean(o["raw_quote"])
        m = o["minute"]
        return f"min {m}: {q}" if m not in _NULLS else q

    wb = Workbook()

    # ── Resumen ──
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Informe de base de datos — Scouting"
    ws["A1"].font = Font(bold=True, size=16, color=GREEN)
    ws["A2"] = "Skip Central Sports · Bot de Scouting"
    ws["A2"].font = Font(italic=True, color="555555")
    days = sorted(s["created_at"][:10] for s in sessions) or ["—"]
    named = [p for p in prospects if _clean(p["name"])]
    rated = [p for p in prospects if _float(p["latest_rating"]) is not None]
    summary_rows = [
        ("Scout", scouts[0]["name"] if scouts else "—"),
        ("Período", f"{days[0]} → {days[-1]}"),
        ("Partidos observados", len(sessions)),
        ("Observaciones totales", len(observations)),
        ("Jugadores con nombre", len(named)),
        ("Perfiles temporales (solo número/foto)", sum(1 for p in prospects if _bool(p["is_temporary"]))),
        ("Jugadores valorados (1–5)", len(rated)),
        ("Notas de equipo", len(team_notes)),
        ("Fecha del informe", report_date),
    ]
    r = 4
    for k, v in summary_rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Jugadores mejor valorados").font = HDR_FONT
    r += 1
    head = ["Jugador", "Club", "Valoración", "Decisión", "Nº obs."]
    for c, h in enumerate(head, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, len(head))
    r += 1
    for p in sorted(rated, key=lambda p: -(_float(p["latest_rating"]) or 0)):
        dec = p["decision_status"] or decision_for_rating(p["latest_rating"]) or "—"
        ws.cell(row=r, column=1, value=p["name"] or "(sin nombre)")
        ws.cell(row=r, column=2, value=p["team"])
        ws.cell(row=r, column=3, value=_float(p["latest_rating"]))
        ws.cell(row=r, column=4, value=dec)
        ws.cell(row=r, column=5, value=len(obs_by_prosp.get(p["id"], [])))
        r += 1
    _widths(ws, [34, 20, 12, 16, 8])

    # ── Partidos ──
    ws = wb.create_sheet("Partidos")
    head = ["#", "Fecha", "Local", "Visitante", "Estado", "Nº observaciones", "Notas equipo"]
    ws.append(head)
    _style_header_row(ws, 1, len(head))
    obs_count: dict[str, int] = defaultdict(int)
    tn_count: dict[str, int] = defaultdict(int)
    for o in observations:
        obs_count[o["session_id"]] += 1
        if _bool(o["is_team_note"]):
            tn_count[o["session_id"]] += 1
    for n, s in enumerate(sorted(sessions, key=lambda s: s["created_at"]), 1):
        ws.append([n, s["created_at"][:10], s["home_team"], s["away_team"],
                   "activo" if s["state"] == "active" else "finalizado",
                   obs_count[s["id"]], tn_count[s["id"]]])
    _widths(ws, [4, 12, 22, 22, 12, 16, 12])

    # ── Jugadores (uno por fila) ──
    ws = wb.create_sheet("Jugadores")
    head = ["Jugador", "Club", "Posición", "Edad", "Estatura (cm)", "Valoración",
            "Decisión", "Nº obs.", "Observaciones agrupadas (con minuto)"]
    ws.append(head)
    _style_header_row(ws, 1, len(head))
    report_prospects = [p for p in prospects if _clean(p["name"]) or obs_by_prosp.get(p["id"])]
    for p in sorted(report_prospects,
                    key=lambda p: (-(_float(p["latest_rating"]) or 0), p["team"] or "", p["name"] or "")):
        grouped = " | ".join(obs_line(o) for o in sorted(obs_by_prosp.get(p["id"], []),
                                                         key=lambda o: _int(o["id"]) or 0))
        dec = p["decision_status"] or decision_for_rating(p["latest_rating"]) or ""
        ws.append([_clean(p["name"]) or "(sin nombre)", p["team"], p["position"],
                   _int(p["age"]), _int(p["height_cm"]), _float(p["latest_rating"]), dec,
                   len(obs_by_prosp.get(p["id"], [])), grouped])
    for row in ws.iter_rows(min_row=2):
        row[-1].alignment = WRAP
        for cell in row[:-1]:
            cell.alignment = TOP
    _widths(ws, [26, 18, 14, 6, 12, 11, 16, 8, 90])

    # ── Observaciones (log completo) ──
    ws = wb.create_sheet("Observaciones")
    head = ["Partido", "Fecha", "Minuto", "Jugador / Nº", "Club", "Fuente", "Sust.", "Valoración", "Nota"]
    ws.append(head)
    _style_header_row(ws, 1, len(head))
    src_map = {"text": "texto", "voice": "voz", "photo": "foto"}
    for o in sorted(observations, key=lambda o: (_int(o["session_id"]) or 0, _int(o["id"]) or 0)):
        if _bool(o["is_team_note"]):
            continue
        s = sess_by_id.get(o["session_id"], {})
        who = o["player_name"] or (f"#{o['player_number']}" if o["player_number"] not in _NULLS else "")
        ws.append([match_label(o["session_id"]), (s.get("created_at") or "")[:10], _int(o["minute"]),
                   who, o["team"], src_map.get(o["source"], o["source"]),
                   "sí" if _bool(o["is_substitution"]) else "", _float(o["rating"]), _clean(o["raw_quote"])])
    for row in ws.iter_rows(min_row=2):
        row[-1].alignment = WRAP
    _widths(ws, [26, 12, 8, 20, 16, 8, 6, 11, 70])

    # ── Notas equipo ──
    ws = wb.create_sheet("Notas equipo")
    head = ["Partido", "Fecha", "Minuto", "Club", "Nota táctica"]
    ws.append(head)
    _style_header_row(ws, 1, len(head))
    for o in sorted(team_notes, key=lambda o: (_int(o["session_id"]) or 0, _int(o["id"]) or 0)):
        s = sess_by_id.get(o["session_id"], {})
        ws.append([match_label(o["session_id"]), (s.get("created_at") or "")[:10],
                   _int(o["minute"]), o["team"], _clean(o["raw_quote"])])
    for row in ws.iter_rows(min_row=2):
        row[-1].alignment = WRAP
    _widths(ws, [26, 12, 8, 18, 80])

    for name in ("Partidos", "Jugadores", "Observaciones", "Notas equipo"):
        wb[name].freeze_panes = "A2"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return {
        "sessions": len(sessions),
        "observations": len(observations),
        "players": len(report_prospects),
        "team_notes": len(team_notes),
    }


def _filter_chat(data: dict[str, list[dict]], chat_id: str) -> dict[str, list[dict]]:
    keep_sessions = {s["id"] for s in data["sessions"] if s["agent_chat_id"] == chat_id}
    return {
        "sessions": [s for s in data["sessions"] if s["agent_chat_id"] == chat_id],
        "scout_profiles": [s for s in data["scout_profiles"] if s["agent_chat_id"] == chat_id],
        "prospects": [p for p in data["prospects"] if p["agent_chat_id"] == chat_id],
        "observations": [o for o in data["observations"] if o["session_id"] in keep_sessions],
    }


def main() -> None:
    repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    today = date.today().isoformat()
    ap = argparse.ArgumentParser(description="Generate the Excel scouting report.")
    ap.add_argument("--source", choices=("render", "csv"), default="render")
    ap.add_argument(
        "--db-id",
        default=None,
        help=f"Render Postgres id (render source); defaults to ${DB_ID_VAR}",
    )
    ap.add_argument("--csv-dir", help="Folder of table CSVs (csv source)")
    ap.add_argument("--chat-id", help="Restrict the report to one Telegram account")
    ap.add_argument("--date", default=today, help="Report date label (default: today)")
    ap.add_argument("--out", default=os.path.join(repo_root, "reports", f"informe-scouting-{today}.xlsx"))
    args = ap.parse_args()

    if args.source == "csv":
        if not args.csv_dir:
            sys.exit("--source csv requires --csv-dir")
        data = load_from_csv(args.csv_dir)
    else:
        data = load_from_render(args.db_id)

    if args.chat_id:
        data = _filter_chat(data, args.chat_id)

    stats = build(data, args.out, args.date)
    print(f"Saved: {args.out}")
    print("Sheets: Resumen | Partidos({sessions}) | Jugadores({players}) | "
          "Observaciones({obs}) | Notas equipo({team_notes})".format(
              obs=stats["observations"] - stats["team_notes"], **stats))


if __name__ == "__main__":
    main()
