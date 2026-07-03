"""Report generation.

Pure functions over hydrated ORM instances (no Telegram/DB), so they're trivially
unit-testable:
  - build_summary(session): short Telegram summary of a finished match.
  - build_workbook(session): the match report — an editable .xlsx workbook with
    Local / Visitante / Notas equipo sheets, ONE row per player (grouped).
  - build_csv(session): flat CSV fallback — one row per observation.
  - build_markdown(session): full markdown match report (used by the REST API).
  - build_player_report(prospect, observations, summary): the accumulated,
    cross-match player report (raw history + AI summary + bio/decision header).

Observation-first: observations carry their own identity snapshot (team, player
name/number/position) and link to a cross-match Prospect. No sentiment/skill —
evaluation is the manual rating only.
"""

from __future__ import annotations

import csv
import io
from collections import defaultdict

from .models import Observation, Prospect, Session, decision_for_rating


def _opponent(session: Session, team: str | None) -> str:
    """The other team in 'A vs B' relative to `team` (blank if unknown)."""
    if not team:
        return ""
    from .taxonomy import normalize_name

    nt = normalize_name(team)
    if nt == normalize_name(session.home_team):
        return session.away_team
    if nt == normalize_name(session.away_team):
        return session.home_team
    return ""


def _match_label(session: Session) -> str:
    return f"{session.home_team} vs {session.away_team}"


def _fmt_date(value) -> str:
    if value is None:
        return ""
    try:
        return value.strftime("%Y-%m-%d")
    except AttributeError:
        return str(value)


# ── Telegram summary ──────────────────────────────────────────────────────────
def build_summary(session: Session) -> str:
    """Short summary shown directly in the Telegram chat after /finalizar."""
    player_obs = [o for o in session.observations if not o.is_team_note]
    team_notes = [o for o in session.observations if o.is_team_note]

    lines = [f"📋 *Informe del partido*: {_match_label(session)}"]
    if session.label:
        lines.append(f"_{session.label}_")
    lines.append(f"Observaciones registradas: *{len(session.observations)}*")

    # Most-noted players, grouped by their identity snapshot (name+team).
    by_player: dict[tuple[str, str], list[Observation]] = defaultdict(list)
    for o in player_obs:
        key = (o.player_name or f"#{o.player_number}" or "?", o.team or "")
        by_player[key].append(o)
    if by_player:
        lines.append("")
        lines.append("*Jugadores más observados*")
        ranked = sorted(by_player.items(), key=lambda kv: len(kv[1]), reverse=True)
        for (name, team), obs in ranked[:5]:
            team_str = f" ({team})" if team else ""
            lines.append(f"• {name}{team_str} — {len(obs)} obs.")

    if team_notes:
        lines.append("")
        lines.append(f"📝 Notas de equipo: {len(team_notes)}")

    lines.append("")
    lines.append("Informe completo adjunto como archivo. 📎")
    return "\n".join(lines)


# ── Match CSV (one row per observation) ─────────────────────────────────────────
_CSV_COLUMNS = [
    "Match",
    "Date",
    "Team",
    "Opponent",
    "Player name",
    "Number",
    "Position",
    "Age",
    "Height",
    "Minute",
    "Observation",
    "Source",
    "Manual rating",
    "Scout",
    "Created at",
]


def build_csv(session: Session) -> bytes:
    """Match report as CSV bytes — one row per observation. UTF-8 with BOM so
    Excel renders Spanish accents. Prospect bio (age/height) and the prospect's
    latest rating fill in when the observation itself doesn't carry them."""
    prospects = {p.id: p for p in _prospects_of(session)}
    match = _match_label(session)
    date = _fmt_date(session.match_date or session.created_at)
    scout = session.scout_name or str(session.agent_chat_id)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_CSV_COLUMNS)
    for o in session.observations:
        p = prospects.get(o.prospect_id) if o.prospect_id else None
        rating = o.rating if o.rating is not None else (p.latest_rating if p else None)
        writer.writerow(
            [
                match,
                date,
                o.team or "",
                _opponent(session, o.team),
                o.player_name or (p.name if p else ""),
                o.player_number if o.player_number is not None else "",
                o.player_position or (p.position if p else ""),
                p.age if p and p.age is not None else "",
                p.height_cm if p and p.height_cm is not None else "",
                o.minute if o.minute is not None else "",
                o.raw_quote,
                o.source or "",
                f"{rating:g}" if rating is not None else "",
                scout,
                o.created_at.isoformat() if o.created_at else "",
            ]
        )
    return buf.getvalue().encode("utf-8-sig")


def _prospects_of(session: Session) -> list[Prospect]:
    """The prospects referenced by this session's observations (prefetched if
    available; falls back to an empty list when not loaded)."""
    seen: dict[int, Prospect] = {}
    for o in session.observations:
        pr = getattr(o, "prospect", None)
        if pr is not None:
            seen[pr.id] = pr
    return list(seen.values())


# ── Excel workbook (one row per player, 3 sheets) ────────────────────────────────
_XLSX_COLUMNS = [
    "Match",
    "Fecha",
    "Local",
    "Visitante",
    "Club jugador",
    "Nombre jugador",
    "Número",
    "Edad",
    "Estatura",
    "Observaciones agrupadas",
    "Valoración final",
    "Decisión",
]
_TEAM_NOTE_COLUMNS = ["Match", "Fecha", "Equipo", "Nota", "Minuto"]


def _grouped_observations(obs: list[Observation]) -> str:
    """Join a player's observations into one cell: 'Gana duelo min 12. Gol min 44.'

    Each observation is its raw quote (capitalized) with its match minute appended
    when known. Ordered as stored (chronological)."""
    parts: list[str] = []
    for o in sorted(obs, key=lambda x: (x.minute if x.minute is not None else 1_000_000, x.id)):
        quote = (o.raw_quote or "").strip().rstrip(".")
        if not quote:
            continue
        quote = quote[0].upper() + quote[1:]
        if o.minute is not None:
            quote = f"{quote} min {o.minute}"
        parts.append(quote + ".")
    return " ".join(parts)


def _player_side(session: Session, team: str | None) -> str | None:
    """'Local'/'Visitante' for a player's team, or None if it matches neither."""
    from .taxonomy import normalize_name

    if not team:
        return None
    nt = normalize_name(team)
    if nt == normalize_name(session.home_team):
        return "Local"
    if nt == normalize_name(session.away_team):
        return "Visitante"
    return None


def _player_rows(session: Session) -> dict[str, list[list]]:
    """Build the per-player rows, keyed by sheet name ('Local'/'Visitante'/'Sin
    equipo'). One row per grouped player, matching _XLSX_COLUMNS."""
    prospects = {p.id: p for p in _prospects_of(session)}
    match = _match_label(session)
    date = _fmt_date(session.match_date or session.created_at)

    # Group player observations by prospect (fallback: name+team snapshot).
    groups: dict[object, list[Observation]] = defaultdict(list)
    for o in session.observations:
        if o.is_team_note:
            continue
        key = o.prospect_id if o.prospect_id else (o.player_name or "", o.team or "")
        groups[key].append(o)

    rows: dict[str, list[list]] = {"Local": [], "Visitante": [], "Sin equipo": []}
    for key, obs in groups.items():
        p = prospects.get(key) if isinstance(key, int) else None
        # Identity: prefer the prospect, else the latest snapshot.
        name = (p.name if p and p.name else "") or _latest(obs, "player_name") or ""
        team = (p.team if p else None) or _latest(obs, "team")
        number = _latest(obs, "player_number")
        # A group with neither a name nor a number is team-level chatter, not a
        # player — skip it (team notes have their own sheet).
        if not name and number is None:
            continue
        age = p.age if p and p.age is not None else None
        height = p.height_cm if p and p.height_cm is not None else None
        rating = (p.latest_rating if p and p.latest_rating is not None
                  else _latest(obs, "rating"))
        decision = (p.decision_status if p and p.decision_status
                    else decision_for_rating(rating)) or ""

        side = _player_side(session, team) or "Sin equipo"
        rows[side].append(
            [
                match,
                date,
                session.home_team,
                session.away_team,
                team or "",
                name,
                number if number is not None else "",
                age if age is not None else "",
                height if height is not None else "",
                _grouped_observations(obs),
                f"{rating:g}" if rating is not None else "",
                decision,
            ]
        )
    return rows


def _latest(obs: list[Observation], attr: str):
    """The most recent non-null value of `attr` across a player's observations."""
    value = None
    for o in sorted(obs, key=lambda x: x.id):
        v = getattr(o, attr, None)
        if v is not None:
            value = v
    return value


def build_workbook(session: Session) -> bytes:
    """The match report as an editable .xlsx: Local, Visitante, Notas equipo.

    One row per player (grouped observations with minutes), plus a team-notes
    sheet. Requires openpyxl (a declared dependency)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    rows_by_side = _player_rows(session)
    # "Sin equipo" players (team never resolved) go on the Local sheet so nothing
    # is silently dropped from the report.
    local_rows = rows_by_side["Local"] + rows_by_side["Sin equipo"]

    header_font = Font(bold=True)

    def _fill(ws, columns: list[str], data: list[list]) -> None:
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
        for row in data:
            ws.append(row)
        # Reasonable column widths.
        for i, _ in enumerate(columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 20

    ws_local = wb.active
    ws_local.title = "Local"
    _fill(ws_local, _XLSX_COLUMNS, local_rows)

    ws_away = wb.create_sheet("Visitante")
    _fill(ws_away, _XLSX_COLUMNS, rows_by_side["Visitante"])

    ws_team = wb.create_sheet("Notas equipo")
    team_rows = [
        [
            _match_label(session),
            _fmt_date(session.match_date or session.created_at),
            o.team or "",
            o.raw_quote,
            o.minute if o.minute is not None else "",
        ]
        for o in session.observations
        if o.is_team_note
    ]
    _fill(ws_team, _TEAM_NOTE_COLUMNS, team_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Cumulative historical workbook (all prospects, all matches) ──────────────────
_HISTORY_COLUMNS = [
    "Nombre jugador",
    "Club",
    "Posición",
    "Edad",
    "Estatura",
    "Partidos observados",
    "Fechas",
    "Observaciones",
    "Valoración final",
    "Decisión",
    "Resumen global",
]


def build_historical_workbook(
    rows: list[tuple[Prospect, list[Observation], str]],
) -> bytes:
    """The cumulative scouting database export: one sheet, one row per prospect,
    aggregating every match they were observed in. `rows` is what
    ScoutingService.historical_report returns. Requires openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico"
    ws.append(_HISTORY_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for prospect, observations, summary in rows:
        matches: dict[str, list[Observation]] = defaultdict(list)
        dates: list[str] = []
        for o in observations:
            session = getattr(o, "session", None)
            label = _match_label(session) if session else "?"
            matches[label].append(o)
            d = _fmt_date(o.created_at)
            if d and d not in dates:
                dates.append(d)
        # Observations grouped per match: "Millonarios vs América: Gol min 44. …"
        blocks = [
            f"{label}: {_grouped_observations(obs)}" for label, obs in matches.items()
        ]
        rating = prospect.latest_rating
        decision = prospect.decision_status or decision_for_rating(rating) or ""
        ws.append(
            [
                prospect.name or "",
                prospect.team or "",
                prospect.position or "",
                prospect.age if prospect.age is not None else "",
                prospect.height_cm if prospect.height_cm is not None else "",
                len(matches),
                ", ".join(dates),
                "\n".join(blocks),
                f"{rating:g}" if rating is not None else "",
                decision,
                summary,
            ]
        )
    for i, _ in enumerate(_HISTORY_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Markdown match report (REST API) ────────────────────────────────────────────
def build_markdown(session: Session) -> str:
    """Full markdown match report (served by the REST /report endpoint)."""
    out: list[str] = [f"# Informe del partido — {_match_label(session)}"]
    if session.label:
        out.append(f"*{session.label}*")
    out.append("")
    out.append(f"- Fecha: {_fmt_date(session.match_date or session.created_at)}")
    if session.competition:
        out.append(f"- Competición: {session.competition}")
    out.append(f"- Total de observaciones: {len(session.observations)}")
    out.append("")

    player_obs = [o for o in session.observations if not o.is_team_note]
    team_notes = [o for o in session.observations if o.is_team_note]

    out.append("## Jugadores")
    if not player_obs:
        out.append("_Sin observaciones de jugadores._")
    else:
        by_player: dict[tuple[str, str], list[Observation]] = defaultdict(list)
        for o in player_obs:
            key = (o.player_name or f"#{o.player_number}" or "?", o.team or "")
            by_player[key].append(o)
        for (name, team), obs in by_player.items():
            team_str = f" — {team}" if team else ""
            out.append(f"### {name}{team_str}")
            for o in obs:
                rating = f" (valoración {o.rating:g})" if o.rating is not None else ""
                out.append(f'- "{o.raw_quote}"{rating}')
            out.append("")

    out.append("## Notas de equipo")
    if team_notes:
        for o in team_notes:
            prefix = f"**{o.team}**: " if o.team else ""
            out.append(f"- {prefix}{o.raw_quote}")
    else:
        out.append("_Ninguna._")
    out.append("")
    return "\n".join(out)


# ── Accumulated cross-match player report ───────────────────────────────────────
def build_player_report(
    prospect: Prospect, observations: list[Observation], summary: str
) -> str:
    """Telegram-formatted accumulated player report: bio/decision header, raw
    cross-match observation history, then the AI-generated summary."""
    name = prospect.name or "Jugador sin nombre"
    lines = [f"👤 *{name}*"]
    bio = []
    if prospect.team:
        bio.append(prospect.team)
    if prospect.position:
        bio.append(prospect.position)
    if prospect.age is not None:
        bio.append(f"{prospect.age} años")
    if prospect.height_cm is not None:
        bio.append(f"{prospect.height_cm} cm")
    if bio:
        lines.append(" · ".join(bio))
    if prospect.latest_rating is not None:
        lines.append(f"Valoración actual: *{prospect.latest_rating:g}*")
    lines.append(f"Decisión: *{prospect.decision_status or 'Pendiente'}*")
    lines.append("")

    lines.append(f"*Historial* ({len(observations)} observaciones)")
    for o in observations:
        session = getattr(o, "session", None)
        match = _match_label(session) if session else ""
        date = _fmt_date(o.created_at)
        rating = f" — valoración {o.rating:g}" if o.rating is not None else ""
        src = f" [{o.source}]" if o.source else ""
        prefix = f"_{date} · {match}_: " if match else f"_{date}_: "
        lines.append(f'• {prefix}"{o.raw_quote}"{rating}{src}')
    lines.append("")

    lines.append("*Resumen*")
    lines.append(summary)
    return "\n".join(lines)
