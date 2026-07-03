"""Match workbook (.xlsx, grouped) + CSV fallback + cross-match player report."""

import csv
import io

import pytest

from scouting_bot.report import (
    _CSV_COLUMNS,
    _XLSX_COLUMNS,
    build_csv,
    build_player_report,
    build_summary,
    build_workbook,
)


async def _capture(service, sess, text, source="text"):
    for r in await service.capture_notes(sess, text, source=source):
        if r.needs_team_choice:
            await service.resolve_team_choice(
                sess, r.classified, r.team_candidates[0], minute=r.minute
            )


def _sheets(content: bytes) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content))
    return {ws.title: [list(r) for r in ws.iter_rows(values_only=True)] for ws in wb}


@pytest.mark.asyncio
async def test_csv_columns_and_opponent(service):
    await service.storage.set_scout_name(1, "Rafa")
    sess, _ = await service.start_session(1, "Millonarios", "América", None)
    await _capture(service, sess, "Castro de América buen pase valoración 4")
    await _capture(service, sess, "América presiona alto")  # team note
    ended = await service.end_session(sess)

    raw = build_csv(ended)
    rows = list(csv.reader(io.StringIO(raw.decode("utf-8-sig"))))
    assert rows[0] == _CSV_COLUMNS
    data = rows[1:]

    col = {name: i for i, name in enumerate(_CSV_COLUMNS)}
    castro = next(r for r in data if "Castro" in r[col["Observation"]])
    assert castro[col["Match"]] == "Millonarios vs América"
    assert castro[col["Team"]] == "América"
    assert castro[col["Opponent"]] == "Millonarios"  # derived
    assert castro[col["Player name"]] == "Castro"
    assert castro[col["Manual rating"]] == "4"
    assert castro[col["Scout"]] == "Rafa"
    assert castro[col["Source"]] == "text"


@pytest.mark.asyncio
async def test_csv_handles_empty_session(service):
    sess, _ = await service.start_session(1, "A", "B", None)
    ended = await service.end_session(sess)
    raw = build_csv(ended)
    rows = list(csv.reader(io.StringIO(raw.decode("utf-8-sig"))))
    assert rows[0] == _CSV_COLUMNS
    assert len(rows) == 1  # header only


# ── xlsx workbook (one row per player, 3 sheets, grouped observations) ──────────
@pytest.mark.asyncio
async def test_workbook_three_sheets_and_grouped_player(service):
    await service.storage.set_scout_name(1, "Rafa")
    sess, _ = await service.start_session(1, "Millonarios", "América", None)
    await service.start_first_half(sess)
    # Same player, several observations → ONE grouped row on the Visitante sheet.
    await _capture(service, sess, "Castro de América gana el duelo aéreo")
    await _capture(service, sess, "Castro de América buena definición valoración 4")
    await _capture(service, sess, "América presiona muy alto")  # team note
    ended = await service.end_session(sess)

    sheets = _sheets(build_workbook(ended))
    assert set(sheets) == {"Local", "Visitante", "Notas equipo"}

    # Visitante sheet: header + exactly one Castro row.
    away = sheets["Visitante"]
    assert away[0] == _XLSX_COLUMNS
    assert len(away) == 2
    col = {name: i for i, name in enumerate(_XLSX_COLUMNS)}
    row = away[1]
    assert row[col["Nombre jugador"]] == "Castro"
    assert row[col["Club jugador"]] == "América"
    assert row[col["Local"]] == "Millonarios" and row[col["Visitante"]] == "América"
    # Both observations grouped into one cell.
    grouped = row[col["Observaciones agrupadas"]]
    assert "duelo aéreo" in grouped and "definición" in grouped
    # Final rating + auto-decision.
    assert str(row[col["Valoración final"]]) == "4"
    assert row[col["Decisión"]] == "Muy interesante"

    # Team note lands on its own sheet.
    team = sheets["Notas equipo"]
    assert any("presiona muy alto" in str(c) for r in team[1:] for c in r)


@pytest.mark.asyncio
async def test_workbook_local_side_and_minutes(service):
    sess, _ = await service.start_session(1, "Millonarios", "América", None)
    await service.start_first_half(sess)
    await service.resync_clock(sess, 12)  # clock ≈ minute 12
    await _capture(service, sess, "Jordan de Millonarios gol")
    ended = await service.end_session(sess)

    sheets = _sheets(build_workbook(ended))
    col = {name: i for i, name in enumerate(_XLSX_COLUMNS)}
    local = sheets["Local"]
    assert len(local) == 2  # header + Jordan
    grouped = local[1][col["Observaciones agrupadas"]]
    assert "min 12" in grouped  # match minute embedded in the grouped text


@pytest.mark.asyncio
async def test_player_report_spans_matches_with_summary(service):
    chat = 3
    m1, _ = await service.start_session(chat, "América", "Nacional", None)
    await _capture(service, m1, "Castro de América buen pase")
    await service.end_session(m1)
    m2, _ = await service.start_session(chat, "Millonarios", "América", None)
    await _capture(service, m2, "Castro de América gana los duelos")
    await service.end_session(m2)

    result = await service.player_report(chat, "Castro")
    assert not isinstance(result, list)
    prospect, observations, summary = result
    assert len(observations) == 2
    report = build_player_report(prospect, observations, summary)
    assert "Castro" in report
    assert "Historial" in report and "Resumen" in report
    assert "buen pase" in report and "gana los duelos" in report
    # Both matches appear in the history.
    assert "América vs Nacional" in report and "Millonarios vs América" in report


# ── cumulative historical workbook ──────────────────────────────────────────────
@pytest.mark.asyncio
async def test_historical_workbook_aggregates_across_matches(service):
    from scouting_bot.report import _HISTORY_COLUMNS, build_historical_workbook

    chat = 7
    m1, _ = await service.start_session(chat, "América", "Nacional", None)
    await _capture(service, m1, "Castro de América buen pase")
    await service.end_session(m1)
    m2, _ = await service.start_session(chat, "Millonarios", "América", None)
    await _capture(service, m2, "Castro de América gana los duelos valoración 4")
    await service.end_session(m2)

    rows = await service.historical_report(chat)
    content = build_historical_workbook(rows)

    sheets = _sheets(content)
    assert list(sheets) == ["Histórico"]
    data = sheets["Histórico"]
    assert data[0] == _HISTORY_COLUMNS
    col = {name: i for i, name in enumerate(_HISTORY_COLUMNS)}
    castro = next(r for r in data[1:] if r[col["Nombre jugador"]] == "Castro")
    assert castro[col["Partidos observados"]] == 2  # both matches aggregated
    assert str(castro[col["Valoración final"]]) == "4"
    assert castro[col["Decisión"]] == "Muy interesante"
    obs_cell = castro[col["Observaciones"]]
    assert "buen pase" in obs_cell and "gana los duelos" in obs_cell


@pytest.mark.asyncio
async def test_summary_groups_by_player(service):
    sess, _ = await service.start_session(1, "A", "B", None)
    await _capture(service, sess, "Castro de A buen pase")
    await _capture(service, sess, "Castro de A buen control")
    ended = await service.end_session(sess)
    summary = build_summary(ended)
    assert "Castro" in summary
    assert "2 obs" in summary  # two observations rolled up for Castro
